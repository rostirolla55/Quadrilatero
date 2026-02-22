import os
import openpyxl

# Percorsi dei file
excel_path = "codice_quadrilatero.xlsx"
word_relative_path = "Come creare una pagina del sito Paginaxy.docx"

def add_hyperlink_to_excel(sheet, row, col, file_path, bookmark, display_text):
    """Aggiunge un collegamento ipertestuale reale a una cella Excel"""
    cell = sheet.cell(row=row, column=col)
    full_link = f"{file_path}#{bookmark}"
    cell.hyperlink = full_link
    cell.value = display_text
    cell.style = "Hyperlink"

def process_links():
    if not os.path.exists(excel_path):
        print(f"Errore: Il file {excel_path} non è stato trovato.")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active 
    except Exception as e:
        print(f"Errore nell'apertura del file Excel: {e}")
        return
    
    # MAPPING ESTESO (Copre i 25 esistenti + suggerimenti per gli altri)
    mappings = {
        # --- ESISTENTI E VERIFICATI ---
        "esegui_all_update_json.bat": "tit_Caricamento_dei_testi_dinamici_scrip",
        "json_updater.py": "tit_Caricamento_dei_testi_dinamici_scrip",
        "update_json.bat": "tit_Caricamento_dei_testi_dinamici_scrip",
        "update_json.py": "tit_Caricamento_dei_testi_dinamici_scrip",
        "manual_key_updater.bat": "tit_Caricamento_dei_testi_statici_manual",
        "manual_key_updater.py": "tit_Caricamento_dei_testi_statici_manual",
        "update_image.bat": "tit_Caricamento_delle_immagini_con_lo_sc",
        "update_image_sources.py": "tit_Caricamento_delle_immagini_con_lo_sc",
        "update_json_image.bat": "tit_caricamento_in_texts_json_dei_nomi_d",
        "update_json_image.py": "tit_caricamento_in_texts_json_dei_nomi_d",
        "add_chiesasancarlo.bat": "tit_Creazione_dello_script_add_paginaxy_",
        "add_pittoricarracci.bat": "tit_Creazione_dello_script_add_paginaxy_",
        "add_pugliole.bat": "tit_Creazione_dello_script_add_paginaxy_",
        "add_bsmariamaggiore.bat": "tit_Creazione_dello_script_add_paginaxy_",
        "add_cavaticcio.bat": "tit_Creazione_dello_script_add_paginaxy_",
        "add_newpage.bat": "tit_Crea_e_modifica_add_paginaxy_bat_NEW",
        "add_page.py": "tit_Crea_e_modifica_add_paginaxy_bat_NEW",
        "salvag.bat": "tit_Crea_e_modifica_add_paginaxy_bat_NEW",
        "dimissione_procedura.bat": "tit_Dopo_aver_eseguito_add_paginaxy_bat",
        "extract_config_poi.py": "tit_Gestione_del_pois_config_json",
        "load_config_poi.py": "tit_Gestione_del_pois_config_json",
        "sync_nav_from_config.py": "tit_Gestione_del_pois_config_json",
        "extract_nav_to_json.py": "tit_Gestione_della_lista_dei_nav_nei_fil",
        "update_html_from_json.py": "tit_Gestione_della_lista_dei_nav_nei_fil",
        "extract_gps.py": "tit_Identificazione_delle_coordinate_GPS",
        "get_coords_from_img.bat": "tit_Identificazione_delle_coordinate_GPS",

        # --- AGGIUNTI DA CSV (Puntano a sezioni generiche o da creare) ---
        "split_and_update_content.py": "tit_Disamina_del_documento_word_paginaxy",
        "sync_config.py": "tit_Disamina_del_documento_word_paginaxy",
        "convert_page.bat": "tit_Disamina_del_documento_word_paginaxy",
        "update_json_key.py": "tit_Caricamento_dei_testi_dinamici_scrip",
        "go_dimissione.py": "tit_Dopo_aver_eseguito_add_paginaxy_bat",
        "cleanup_fragments.py": "tit_Dopo_aver_eseguito_add_paginaxy_bat",
        
        # --- OBSOLETI O DA DOCUMENTARE (Puntano all'inizio per ora) ---
        "pulizia_html.py": "tit_Disamina_del_materiale_ricevuto",
        "process_all_pages.py": "tit_Disamina_del_materiale_ricevuto",
        "key_synchronization.py": "tit_Caricamento_dei_testi_dinamici_scrip",
        "convert_docx_to_html.bat": "tit_Disamina_del_documento_word_paginaxy",
        "convert_docx_to_html.py": "tit_Disamina_del_documento_word_paginaxy",
        "snapshot_sito.bat": "tit_Crea_e_modifica_add_paginaxy_bat_NEW"
    }

    found_count = 0
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if not cell_val:
            continue
            
        file_name = str(cell_val).strip().lower()
        
        for key, bookmark in mappings.items():
            if file_name == key.lower():
                add_hyperlink_to_excel(
                    ws, 
                    row=row, 
                    col=4, 
                    file_path=word_relative_path, 
                    bookmark=bookmark, 
                    display_text=bookmark
                )
                found_count += 1
                break

    wb.save(excel_path)
    print(f"Processo terminato. {found_count} collegamenti gestiti.")

if __name__ == "__main__":
    process_links()