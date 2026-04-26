import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Creazione del workbook e del foglio
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Flusso Creazione Pagina"

# Definizione degli stili
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
cmd_font = Font(name="Courier New", bold=True, color="006100")
cmd_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Intestazioni
headers = ["Fase", "Attività", "Eseguire in CMD?", "Comando / Azione", "Descrizione e Note"]
ws.append(headers)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment_center
    cell.border = border

# Dati del flusso
data = [
    ["1. PREPARAZIONE", "Backup del sito", "SÌ", "backup PortoReno", "Crea uno snapshot dello stato attuale prima delle modifiche."],
    ["1. PREPARAZIONE", "Raccolta foto GPS", "NO", "Smartphone (GPS ON)", "Scattare foto sul posto e rinominarla 'paginaxy.jpg'."],
    ["1. PREPARAZIONE", "Testi e Immagini", "NO", "Word / Editor", "Preparare [paginaxy]_it.docx con tag [SPLIT_BLOCK] e traduzioni."],
    ["2. CONFIGURAZIONE", "Estrazione Coordinate", "SÌ", "get_coords_from_img.bat", "Estrae Latitudine e Longitudine dalla foto scattata."],
    ["2. CONFIGURAZIONE", "Configurazione POI", "NO", "Editor JSON", "Inserire ID, Titolo e Coordinate in 'pois_config.json'."],
    ["2. CONFIGURAZIONE", "Generazione Script Pagina", "SÌ", "create_newpage.bat [nome_pagina]", "Genera lo script specifico 'add_[nome_pagina].bat'."],
    ["3. CONVERSIONE", "Conversione Contenuti (IT)", "SÌ", "convert_page.bat it [nome_pagina]", "Converte il file Word in frammenti HTML e estrae immagini."],
    ["3. CONVERSIONE", "Conversione Contenuti (altre lingue)", "SÌ", "convert_page.bat [lingua] [nome_pagina]", "Ripetere per 'en', 'es', 'fr' per aggiornare il database testi."],
    ["4. INTEGRAZIONE", "Integrazione nel Sito", "SÌ", "add_[nome_pagina].bat", "Inserisce la pagina nella struttura e aggiorna main.js."],
    ["4. INTEGRAZIONE", "Sincronizzazione Menu", "SÌ", "python sync_nav_from_config.py", "Aggiorna la lista di navigazione in tutti i file HTML."],
    ["4. INTEGRAZIONE", "Aggiornamento Testi Dinamici", "SÌ", "update_json.bat", "Sincronizza i testi estratti con il database multilingua."],
    ["4. INTEGRAZIONE", "Testi Statici e Orari", "SÌ", "manual_key_updater.bat", "Inserimento manuale di chiavi specifiche (orari, link, fonti)."],
    ["5. AUDIO", "Produzione Audio", "NO", "Audacity / Gemini", "Registrare l'audio e caricarlo in Assets/Audio/[lingua]/."]
]

for row in data:
    ws.append(row)

# Applicazione stili alle righe di dati
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = border
        if cell.column == 3: # Colonna "Eseguire in CMD?"
            cell.alignment = alignment_center
        else:
            cell.alignment = alignment_left
        
        if cell.column == 4 and row[2].value == "SÌ": # Colonna Comandi
            cell.font = cmd_font
            cell.fill = cmd_fill

# Regolazione larghezza colonne
column_widths = [18, 30, 15, 35, 50]
for i, width in enumerate(column_widths):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width

# Salvataggio
file_path = "piano_operativo_paginaxy.xlsx"
wb.save(file_path)cd